# excel_formatter.py
from __future__ import annotations

import re
from typing import Dict, Any, Tuple, List
import openpyxl
import numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Excel number formats (edit as desired)
FMT = {
    "Currency": "$#,##0.00",
    "Currency2": "$#,##0.00",
    "Percentage": "0.00%",
    "Percentage4": "0.0000%",
    "Integer": "#,##0",
    "Date": "yyyy-mm-dd",
}

RANGE_RE = re.compile(r"^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$")


def _autosize_columns(ws, max_width: int = 60, scan_rows: int = 2000):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[: min(len(col), scan_rows)]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)


def _set_col_format(ws, col_idx: int, number_format: str, start_row: int = 2):
    col_letter = get_column_letter(col_idx)
    for row in range(start_row, ws.max_row + 1):
        ws[f"{col_letter}{row}"].number_format = number_format


def _set_range_format(ws, cell_range: str, number_format: str):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = number_format

def format_backlog_dynamic(ws, currency_format: str = FMT["Currency"]):
    """
    Rolling 18 Backlog sheet:
      - leave header row alone
      - detect metadata cols by header name
      - format all remaining cells (rows 2+) as currency
    """
    if ws.max_row < 2 or ws.max_column < 2:
        return

    header = [c.value for c in ws[1]]
    meta = {"Job", "JobName", "JobSearch", "Division", "MarketSegment"}

    # first column index (1-based) that is NOT metadata
    first_money_col = None
    for idx, h in enumerate(header, start=1):
        if h is None:
            continue
        if str(h).strip() not in meta:
            first_money_col = idx
            break

    # fallback: old behavior if something unexpected
    if first_money_col is None:
        first_money_col = 6  # assumes 5 metadata cols

    for row in range(2, ws.max_row + 1):
        for col in range(first_money_col, ws.max_column + 1):
            ws.cell(row=row, column=col).number_format = currency_format

def add_backlog_totals_row(ws, label_col: int = 1, currency_format: str = FMT["Currency"]):
    if ws.max_row < 2 or ws.max_column < 2:
        return

    header = [c.value for c in ws[1]]
    meta = {"Job", "JobName", "JobSearch", "Division", "MarketSegment"}

    first_money_col = None
    for idx, h in enumerate(header, start=1):
        if h is None:
            continue
        if str(h).strip() not in meta:
            first_money_col = idx
            break
    if first_money_col is None:
        first_money_col = 6

    # find last real data row by scanning col A (Job)
    last_data_row = None
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=label_col).value
        if v is not None and str(v).strip() != "":
            if str(v).strip().upper() == "TOTAL":
                return
            last_data_row = r
            break
    if last_data_row is None:
        return

    total_row = last_data_row + 1
    ws.cell(row=total_row, column=label_col).value = "TOTAL"

    for col in range(first_money_col, ws.max_column + 1):
        s = 0.0
        for r in range(2, last_data_row + 1):
            val = ws.cell(row=r, column=col).value
            if isinstance(val, numbers.Number):
                s += float(val)

        c = ws.cell(row=total_row, column=col)
        c.value = s
        c.number_format = currency_format

    bold = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=total_row, column=col).font = bold

def add_totals_row_fixed_cols(
    ws,
    label: str = "TOTAL",
    label_col: int = 1,
    first_sum_col: int = 6,   # F
    last_sum_col: int = 10,   # J
    currency_format: str = FMT["Currency"]
):
    if ws.max_row < 2 or ws.max_column < last_sum_col:
        return

    # find last real data row by scanning label col
    last_data_row = None
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=label_col).value
        if v is not None and str(v).strip() != "":
            if str(v).strip().upper() == label:
                return
            last_data_row = r
            break
    if last_data_row is None:
        return

    total_row = last_data_row + 1
    ws.cell(row=total_row, column=label_col).value = label

    for col in range(first_sum_col, last_sum_col + 1):
        s = 0.0
        for r in range(2, last_data_row + 1):
            val = ws.cell(row=r, column=col).value
            if isinstance(val, numbers.Number):
                s += float(val)

        c = ws.cell(row=total_row, column=col)
        c.value = s
        c.number_format = currency_format

    bold = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=total_row, column=col).font = bold

def load_format_spec_from_xlsx(spec_path: str) -> Dict[str, List[Tuple[str, str]]]:
    """
    Reads a 3-column sheet:
      Sheet | Cell/Column Name | Format

    Returns:
      {
        "WIP": [("earned_revenue","Currency"), ("cost_percent_complete","Percentage"), ...],
        "Rolling 18 Backlog": [("D2:W16","Currency"), ...],
        ...
      }
    """
    wb = openpyxl.load_workbook(spec_path, data_only=True)
    ws = wb.active

    spec: Dict[str, List[Tuple[str, str]]] = {}

    # assume header row 1
    for r in range(2, ws.max_row + 1):
        sheet = ws.cell(r, 1).value
        target = ws.cell(r, 2).value
        fmt = ws.cell(r, 3).value

        if not sheet or not target or not fmt:
            continue

        sheet = str(sheet).strip()
        target = str(target).strip()
        fmt = str(fmt).strip()

        spec.setdefault(sheet, []).append((target, fmt))

    return spec


def apply_formats(workbook: openpyxl.Workbook, spec: Dict[str, List[Tuple[str, str]]], autosize: bool = True):
    """
    Applies formatting to workbook, driven by spec.
    Supports:
      - target = "A1:D10" (range)
      - target = "ColumnName" (header match on row 1)
    """
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True

    if "Rolling 18 Backlog" in workbook.sheetnames:
        ws = workbook["Rolling 18 Backlog"]

        # add totals first, then format everything D: to end (rows 2+)
        add_backlog_totals_row(ws)
        format_backlog_dynamic(ws, currency_format=FMT["Currency"])

        if autosize:
            _autosize_columns(ws)

    if "Job Revenue (WIP MoM Latest)" in workbook.sheetnames:
        ws = workbook["Job Revenue (WIP MoM Latest)"]
        add_totals_row_fixed_cols(ws, first_sum_col=6, last_sum_col=10, currency_format=FMT["Currency"])
        if autosize:
            _autosize_columns(ws)

    for sheet_name, rules in spec.items():
        if sheet_name not in workbook.sheetnames:
            continue

        if sheet_name.strip().lower() == "rolling 18 backlog":
            continue


        ws = workbook[sheet_name]
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []

        for target, fmt_key in rules:
            number_format = FMT.get(fmt_key, fmt_key)  # allow raw excel format strings too

            if RANGE_RE.match(target):
                # range like D2:W16
                _set_range_format(ws, target, number_format)
                continue

            # else interpret as column name
            if target in header:
                col_idx = header.index(target) + 1
                _set_col_format(ws, col_idx, number_format, start_row=2)

        if autosize:
            _autosize_columns(ws)
